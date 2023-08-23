import os.path

from appearence import bcolors
import openpyxl
import pandas as pd


class CollectedReport:
    report_path = ""
    updated_afus = []

    def __init__(self, report_path):
        self.report_path = report_path
        self.updated_afus = self.__getCollectedAFUs()

    def __getCollectedAFUs(self):
        wb = openpyxl.load_workbook(self.report_path)
        sheet = wb.active
        updated_afus = []
        for rows in sheet.iter_rows(min_row=2, max_row=2500, min_col=5, max_col=17):
            col = 5
            updated_row = ()
            stop = False
            for cell in rows:
                if col == 5 and cell.value is None:
                    stop = True
                    break
                if col == 5 or col == 17:
                    value = cell.value
                    if col == 17:
                        value = value.strftime("%#m/%#d/%Y")
                    updated_row += (value,)
                col += 1
            if stop:
                break
            updated_afus.append(updated_row)
        return updated_afus

    def dateEntered(self, child_id):
        for tupleRow in self.updated_afus:
            if tupleRow[0] == child_id:
                return tupleRow[1]
        return ""


def syncWithReport(manual_afu_path):
    report_path = input(f"Enter report path: ")
    if not os.path.exists(report_path):
        input("Incorrect file path! press enter to go back")
        return
    collected_report = CollectedReport(report_path)

    wb = openpyxl.load_workbook(manual_afu_path)
    sheet = wb.active

    id_col = 3
    start_row = 3
    row = start_row
    count = 0
    for rows in sheet.iter_rows(min_row=start_row, max_row=8797, min_col=id_col, max_col=16):
        if str(sheet.cell(row=row, column=17).value) == "None":
            write_date = collected_report.dateEntered(str(sheet.cell(row=row, column=id_col).value))
            if write_date != "":
                sheet.cell(row=row, column=17).value = write_date
                count += 1
                print(f"{count}.{sheet.cell(row=row, column=2).value}:{write_date}")
        row += 1
    wb.save(manual_afu_path)
    input("Manual AFU Synced! Press enter to continue")


def isCorrectManualAfuFile(manual_afu_path, other_manual_afu_path):
    wb_main = openpyxl.load_workbook(manual_afu_path)
    sheet_main = wb_main.active
    wb = openpyxl.load_workbook(other_manual_afu_path)
    sheet = wb.active

    incorrect_file = False

    # make sure both file are of the same format by comparing all column header
    row = 2
    col = 1
    for rows in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=16):
        for cell in rows:
            incorrect_file = cell.value != sheet_main.cell(row=row, column=col).value
            print(cell.value + " <--> " + sheet_main.cell(row=row, column=col).value)
            col += 1
            if incorrect_file:
                break
        row += 1
        col = 1
        if incorrect_file:
            break
    if incorrect_file:
        return False
    return True


def copyFrom(manual_afu_path):
    source_path = input(f"Enter the other manual AFU's path: ")
    if not os.path.exists(source_path):
        input("Path invalid! Press enter key to go back.")
        return
    if isCorrectManualAfuFile(manual_afu_path, source_path):
        print(f"{bcolors.BOLD}Start copying...{bcolors.ENDC}")
        wb_dest = openpyxl.load_workbook(manual_afu_path)
        sheet_dest = wb_dest.active
        wb_source = openpyxl.load_workbook(source_path)
        sheet_source = wb_source.active

        row = 3
        col = 3
        row_copy = 0
        for rows in sheet_dest.iter_rows(min_row=3, max_row=9143, min_col=3, max_col=16):
            skip_row = False
            worth_copy = False
            for cell in rows:
                # Make sure only import if child id of the row matched!
                source_cell = sheet_source.cell(row=row, column=col)
                dest_cell = sheet_dest.cell(row=row, column=col)
                if col == 3:
                    if str(dest_cell.value).strip() != str(source_cell.value).strip():
                        print(f"Row skipped! Diff :{dest_cell.value}<-->{source_cell.value}")
                        skip_row = True
                # copy from column 9 (Grade) to the last column

                if col > 8 and not skip_row:
                    if str(source_cell.value).strip() != "None" and str(
                            dest_cell.value).strip() == "None":
                        dest_cell.value = source_cell.value
                        worth_copy = True
                col += 1
            if (not skip_row) and worth_copy:
                row_copy += 1
            row += 1
            col = 2
        wb_dest.save(manual_afu_path)
        print(f"{bcolors.OKGREEN}Copy content done!{bcolors.ENDC}")
        print(f"{bcolors.BOLD}{row_copy}{bcolors.ENDC} rows copied!")
        input("Press enter to continue")
        return True

    print(f"{bcolors.FAIL}Files are of different format!{bcolors.ENDC}")
    input("Press enter to continue")
    return False

def getLatestAttributeCodes(cursor) -> object:
    # DEFINING STANDARD CODES
    grades = []
    daily_activities = []
    ambitions = []
    favorite_subjects = []
    favorite_plays = []
    personalities = []

    # DEFINE THE REQUIRED ATTRIBUTES
    required_attr_codes = [1, 3, 4, 5, 6, 8]
    grouped_attributes = [[], [], [], [], [], []]

    i = 0
    print("Collecting latest valid attributes values...")
    for attr_code in required_attr_codes:
        cursor.execute(
            f"select CH_PROP_CODE from dbo_MASTERTABLE_CHILD_PROFILE where CH_ATTR_CODE ={attr_code}")
        grouped_attributes[i] = cursor.fetchall()
        if attr_code == 1:
            personalities = [x[0] for x in grouped_attributes[i]]

        elif attr_code == 3:
            favorite_subjects = [x[0] for x in grouped_attributes[i]]

        elif attr_code == 4:
            grades = [x[0] for x in grouped_attributes[i]]

        elif attr_code == 5:
            daily_activities = [x[0] for x in grouped_attributes[i]]

        elif attr_code == 6:
            ambitions = [x[0] for x in grouped_attributes[i]]

        elif attr_code == 8:
            favorite_plays = [x[0] for x in grouped_attributes[i]]

        i += 1

    return {
        "grades": grades,
        "personalities": personalities,
        "daily_activities": daily_activities,
        "favorite_subjects": favorite_subjects,
        "ambitions": ambitions,
        "favorite_plays": favorite_plays
    }


def profileDataAreValid(db_format_profiles, cursor):
    # PERFORM VALIDATION ON ALL FILLED ATTRIBUTE ON MANUAL AFU
    # IN ORDER TO MAKE SURE ALL CODES FOLLOW THE RULE
    correct_attrib_codes = getLatestAttributeCodes(cursor)
    warn = False
    print("Validating attributes... Please wait...(3/4)")
    for row in db_format_profiles:
        failed_attributes = []
        if row[1] not in correct_attrib_codes["grades"]:
            failed_attributes.append(f"Grade {row[1]}")
        # 2:personality, 3: daily_activity, 4: fav_subject, 5: ambition, 6: fav_play
        if row[2] not in correct_attrib_codes["personalities"]:
            failed_attributes.append(f"Personality {row[2]}")
        if row[3] not in correct_attrib_codes["daily_activities"]:
            failed_attributes.append(f"Daily Activity {row[3]}")
        if row[4] not in correct_attrib_codes["favorite_subjects"]:
            failed_attributes.append(f"Favorite Subject {row[4]}")
        if row[5] not in correct_attrib_codes["ambitions"]:
            failed_attributes.append(f"Ambition {row[5]}")
        if row[6] not in correct_attrib_codes["favorite_plays"]:
            failed_attributes.append(f"Favorite Play {row[6]}")

        if len(failed_attributes) > 0:
            warn = True
            print(
                f"{bcolors.FAIL}Incorrect attributes for child {bcolors.WARNING}{int(row[0])}:{bcolors.ENDC}{bcolors.ENDC}")
            print(f"{bcolors.WARNING}{failed_attributes}{bcolors.ENDC}")
    if not warn:
        print(f"{bcolors.OKGREEN}{bcolors.BOLD}All Fields Are VALID!{bcolors.BOLD}{bcolors.ENDC}")
        input("Press enter to go back")
        return True
    return False


def getManualAfuDbFormat(manual_afu_path, school_name):
    print("Reading manual afu content... Please wait...")
    df = pd.read_excel(manual_afu_path, sheet_name='AFU_Detail', skiprows=1)

    # Only read data from a school that are never been entered, and has all field filled
    filtered_records = df[df["Sekolah"] == school_name]
    # Only keep column Child Id, Grade, Daily Activity	Ambition, Favorite Subject, Favorite Play, Personality, Address,
    # Write Date
    filtered_records = filtered_records.filter(
        ['ID Anak', 'Kelas Baru', 'Kepribadian', 'Aktivitas Sepulang Sekolah', 'Pelajaran Kesukaan', 'Cita-cita',
         'Permainan Kesukaan',
         'Date Entered'])
    # Only validate complete rows that not already entered
    filtered_records = filtered_records[filtered_records["Date Entered"].isna()]
    filtered_records = filtered_records[filtered_records["Kelas Baru"].fillna("") != ""]
    filtered_records = filtered_records[filtered_records["Aktivitas Sepulang Sekolah"].fillna("") != ""]

    filtered_records["ID Anak"] = filtered_records["ID Anak"].astype(int).astype(str)
    filtered_records["Kelas Baru"] = filtered_records["Kelas Baru"].fillna(0).astype(int)
    filtered_records["Kepribadian"] = filtered_records["Kepribadian"].fillna(0).astype(int)
    filtered_records["Aktivitas Sepulang Sekolah"] = filtered_records["Aktivitas Sepulang Sekolah"].fillna(
        0).astype(int)
    filtered_records["Pelajaran Kesukaan"] = filtered_records["Pelajaran Kesukaan"].fillna(0).astype(int)
    filtered_records["Cita-cita"] = filtered_records["Cita-cita"].fillna(0).astype(int)
    filtered_records["Permainan Kesukaan"] = filtered_records["Permainan Kesukaan"].fillna(0).astype(int)

    filtered_records["Date Entered"] = filtered_records["Date Entered"].fillna('')

    # convert pandas data frame to array of tuples
    db_format_data = filtered_records.to_records(index=False)
    return db_format_data

