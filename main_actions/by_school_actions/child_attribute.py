import openpyxl
from openpyxl.styles import PatternFill
from main_actions.general_actions.manual_afu import profileDataAreValid, getManualAfuDbFormat
from appearence import bcolors
from models.children import getAllChildren
from datetime import date
from helpers.files import getChildIdsFromPath
from sqlite_data.children import getCommunity, getSchool
import os
import shutil


def extractChildIdsFrom(profile_rows):
    child_ids = []
    for group in profile_rows:
        for row in group:
            child_ids.append(row[0])
    return child_ids


class ChildAttribute:
    def __init__(self, manual_afu_path, school_name, profile_status, cursor):
        self.INCLUDE_ADDRESS = False
        self.DEFAULT_PHOTO_PATH = "D:\\BY_SCHOOL_AFU_PHOTOS\\"

        self.__manual_afu_path = manual_afu_path
        self.__school_name = school_name
        self.__cursor = cursor
        # Indicate whether attribute codes from manual afu (selected school) are all valid (True)
        # or contain non-standard codes (False)
        self.__profiles_are_valid = profile_status
        # Will contain array of tuples converted from pandas data frame from manual afu Excel file
        self.__db_format_profiles = []

        # keep rows from @self.matched_profiles that exist in current selected E-AFU and has photo file,
        # divided to per-prefixes array
        # its format will be [[rows from self.db_format_profiles with prefixes _1 in E-AFU DB],[..._2],[...3],...etc]
        self.__matched_profiles = []
        # keep rows from @self.matched_profiles that NOT EXIST in current selected E-AFU
        # will contain some rows(if any) from @self.db_format_profiles
        self.__non_matched_profiles = []
        self.__has_no_photo_profiles = []

    # Main steps for updating child attributes data on E-AFU
    # return True if whole process complete normally
    def runWholeProcess(self):
        # Step 1: Get profile data from manual afu, filtered by school name
        self.__db_format_profiles = self.__getProfileDataBySchool()

        # Step 2: Validate all profile data if not yet validated
        if not self.__profiles_are_valid:
            self.__profiles_are_valid = self.__validateProfileData()
        if self.__profiles_are_valid:
            # Step 3: Split records of @self.db_format_profiles to @self.matched_profiles,
            # @self.non_matched_profiles, and @self.has_no_photo_profiles
            # @self.matched_profiles contain records that exist in current E-AFU
            # and also has new photo in folder @self.DEFAULT_PHOTO_PATH\\@self.school_name
            # the rest records will go to self.non_matched_profiles
            self.__matched_profiles, \
                self.__non_matched_profiles, \
                self.__has_no_photo_profiles = self.__splitExistNonExistNoPhoto()
            # Run Whole Update Process for self.matched_profiles records
            self.__runMatchedProfileUpdate()
            # Rename manual AFU's school for children in @self.non_matched_profiles
            self.__renameUnmatchedProfilesSchoolName()
            # Put mark on manual AFU's for children in @self.has_no_photo_profiles
            self.__markedNoPhotoRows()
            return True
        else:
            return False

    def dbFormatOfManualAfu(self):
        return self.__db_format_profiles

    # Update a child photo
    def __updateSingleChildPhoto(self, child_id):
        source_path = self.DEFAULT_PHOTO_PATH+self.__school_name

    # Update Children Photos
    def __updatePhotos(self):
        source_path = self.DEFAULT_PHOTO_PATH+self.__school_name
        files = []

        def get_files(source_dir, ext="jpg"):
            print("Deep reading file(s)...")
            for entry in os.listdir(source_dir):
                if os.path.isfile(os.path.join(source_dir, entry)):
                    file_name = os.path.join(source_dir, entry)
                    if file_name.split(".")[-1].lower() == ext.lower():
                        files.append(file_name)

                elif os.path.isdir(os.path.join(source_dir, entry)):
                    sub_dir = os.path.join(source_dir, entry)
                    get_files(sub_dir)

        get_files(source_path, "jpg")
        children_with_photo = []
        # (ch_child_id_, HS_, HS_roll_shot_, HS_filename_)
        raw_data_sources = []
        for file in files:
            child_id = rf"{file}".split("\\")[-1].split("_")[0]
            children_with_photo.append(child_id)
            shoot_number = rf"{file}".split("\\")[-1].split("_")[-1].split(".")[0]

            # prepare tuples for db injection
            HS_roll_shot_ = shoot_number
            HS_filename_ = f"{child_id}_{shoot_number}.JPG"
            HS_ = rf"G:\DataInput\AFU\Photo\NewPhoto\{HS_filename_}"
            photo_record = (child_id, HS_, HS_roll_shot_, HS_filename_)
            raw_data_sources.append(photo_record)

        print(f"children with photo: {len(children_with_photo)}")

        all_children = getAllChildren(self.__cursor)
        filtered_groups = [[], [], [], [], [], [], [], []]
        total_exist = 0
        for child_id in children_with_photo:
            id_exist = False
            count = -1
            for children_group in all_children:
                count += 1
                clean_children_group = [item[0] for item in children_group]
                id_exist = child_id in clean_children_group
                if id_exist:
                    total_exist += 1
                    filtered_groups[count].append(child_id)
                    break
            if not id_exist:
                print(f"id {bcolors.FAIL}{child_id}{bcolors.ENDC} not found!")
        i = 0
        for group in filtered_groups:
            if len(group) > 0:
                print(f"children_{i + 1}: {len(group)}")
            i += 1

        total_idHasPhoto = sum([len(group) for group in filtered_groups])
        print(f"Total Id In DB That Has Photo: {total_idHasPhoto}")

        strange = len(children_with_photo) != total_idHasPhoto

        if strange:
            print(f"{bcolors.FAIL}\n\nTotal photo files is different than total detected Ids!{bcolors.ENDC}")
            print(
                f"{bcolors.BOLD}Difference:{bcolors.ENDC} {bcolors.OKGREEN}{len(files)}{bcolors.ENDC} vs {bcolors.OKGREEN}{total_idHasPhoto}{bcolors.ENDC}")
            print("Looking for photo files that doesn't match in...")
            total_exist = 0
            idx = 0
            for child_id in children_with_photo:
                id_exist = False
                for children_group in all_children:
                    clean_children_group = [item[0] for item in children_group]
                    id_exist = child_id in clean_children_group
                    if id_exist:
                        total_exist += 1
                        break
                if not id_exist:
                    print(f"Id {bcolors.WARNING}{child_id}{bcolors.ENDC} in photo folder didn't match Ids in DB!")
                    print(f"Preparing moving file...")
                    dest_dir = self.DEFAULT_PHOTO_PATH + "\\" + getCommunity(child_id) + " - " + getSchool(child_id)
                    if not os.path.isdir(dest_dir):
                        os.mkdir(dest_dir)

                    s_path = os.path.join(source_path, raw_data_sources[idx][3])
                    dest_path = os.path.join(dest_dir, raw_data_sources[idx][3])
                    shutil.move(s_path, dest_path)
                    print(child_id + " to " + dest_path)
                idx += 1
            print(f"{bcolors.OKGREEN}{total_exist}{bcolors.ENDC} photo has existing id!")
            input("Do you want to continue? [y/n]")
        else:
            ##PROCEED TO UPDATE PHOTOS TO DB
            print(f"{bcolors.OKGREEN}All IDs in Photo Folder and IDs in DB FULLY MATCH!{bcolors.ENDC}")
            print(f"{bcolors.BOLD}Executing update...{bcolors.ENDC}")
            postfix = 0
            for filteredIds in filtered_groups:
                postfix += 1
                for child_id in filteredIds:
                    currentRawRecord = [record for record in raw_data_sources if record[0] == child_id]
                    currentChildData = currentRawRecord[0]
                    print(f"{bcolors.BOLD}Updating record:...{bcolors.ENDC}")
                    print(currentChildData)
                    sql_string = fr" UPDATE CPhoto SET HS_roll_shot_{postfix}='{currentChildData[2]}', HS_filename_{postfix}='{currentChildData[3]}' WHERE ch_child_id_{postfix}='{currentChildData[0]}'"
                    print(sql_string)
                    self.__cursor.execute(sql_string)
                    self.__cursor.commit()
            print(f"{bcolors.OKGREEN}{bcolors.BOLD}DONE UPDATE CHILD PHOTOS!{bcolors.ENDC}{bcolors.ENDC}")

    # Update AFU of records from manual AFU that exist in current E-AFU
    def __runMatchedProfileUpdate(self):
        postfix = 0
        add_notes_option = True
        for suffixed_group in self.__matched_profiles:
            postfix += 1
            # get old data for comparison
            if len(suffixed_group) > 0:
                print(f"Update group-{postfix}")
                for row in suffixed_group:
                    child_id = row[0]

                    old_row = self.__getOldProfile(child_id, postfix)

                    notes = date.today().strftime("%b-%d-%Y")

                    # assigning new profiles
                    new_grade = row[1]
                    new_personality = row[2]
                    new_daily_activity = row[3]
                    new_fav_subject = row[4]
                    new_ambition = row[5]
                    new_fav_play = row[6]

                    # get old profiles for comparisson
                    old_grade = old_row[1]
                    old_personality = old_row[2]
                    old_daily_activity = old_row[3]
                    old_fav_subject = old_row[4]
                    old_ambition = old_row[5]
                    old_fav_play = old_row[6]
                    # Grade changed
                    if new_grade != old_grade:
                        if abs(new_grade - old_grade > 2):
                            print(f"{bcolors.WARNING}Strange grade change for child {row[0]}{bcolors.ENDC}")
                            notes += "_Grade corrected"
                        elif abs(new_grade - old_grade == 2):
                            notes += "_Grade corrected"
                        else:
                            notes += "_Grade"
                    # Personality changed
                    if new_personality != old_personality:
                        notes += "_Personality"
                    # Daily Activity changed
                    if new_daily_activity != old_daily_activity:
                        notes += "_Daily activity"
                    # Favorite subject changed
                    if new_fav_subject != old_fav_subject:
                        notes += "_Favorite subject"
                    # Ambition changed
                    if new_ambition != old_ambition:
                        notes += "_Ambition"
                    # Favorite play changed
                    if new_fav_play != old_fav_play:
                        notes += "_Favorite play"
                    notes += "_new photo"
                    notes += "///"
                    print(f"Notes : {notes}")
                    note_sql = f", notes_{postfix}='{notes}' "
                    if not add_notes_option:
                        note_sql = ""

                    # UPDATE CHILD PROFILE
                    self.__updateChildProfile(child_id, postfix, new_grade, new_personality, new_daily_activity, new_fav_subject, new_ambition, new_fav_play,note_sql)

                    # UPDATE TRACKING TABLES
                    self.__updateTrackingTable(child_id, postfix)

                    # 2-> AFU_House
                    house_id = old_row[7]
                    # 2.1. Set AFU_House completed mark
                    self.__setHouseComplete(house_id)
                    # 2.2 Update Address of AFU_House
                    if self.INCLUDE_ADDRESS:
                        new_address = row[7]
                        if new_address != '' and len(new_address.split(' ')) > 1:
                            self.__updateHouseAddress(house_id, new_address)
                # UPDATE CHILDREN's PHOTO
                self.__updatePhotos()
        # Update manual afu (column date entered) with current date if the column is empty
        self.__updateCollectedDateManualAfu()

    # Get nom of E-AFU rows in self.matched_profiles
    def numManualAfuMatchedEAFURows(self):
        return sum([len(group) for group in self.__matched_profiles])

    def __getOldProfile(self, child_id, postfix):
        # Column order adjusted to mimic the column order of manual afu
        sql_string = f"select ch_child_id_{postfix}, ch_prof_grade_{postfix}, ch_prof_personality_{postfix}, " \
                     f"ch_prof_act_{postfix}, ch_prof_subj_{postfix},ch_prof_ambition_{postfix}, " \
                     f"ch_prof_fav_play_{postfix}, ID from AFU_Child WHERE ch_child_id_{postfix}='" \
                     f"{child_id}' "
        self.__cursor.execute(sql_string)
        result = self.__cursor.fetchall()
        return result[0]

    def __updateChildProfile(self, child_id, postfix, grade, personality, daily_act, fav_subject, ambition, fav_play, note_part_sql):
        sql_update_string = f"UPDATE AFU_Child SET " \
                            f"ch_prof_grade_{postfix}={grade}, " \
                            f"ch_prof_personality_{postfix}={personality}, " \
                            f"ch_prof_act_{postfix}={daily_act}, " \
                            f"ch_prof_subj_{postfix}={fav_subject}, " \
                            f"ch_prof_ambition_{postfix}={ambition}, " \
                            f"ch_prof_fav_play_{postfix}={fav_play} " \
                            f"{note_part_sql} " \
                            f"WHERE ch_child_id_{postfix}='{child_id}'"
        self.__cursor.execute(sql_update_string)
        self.__cursor.commit()
        return True

    def __updateTrackingTable(self, child_id, postfix):
        date_note = date.today().strftime("%#m/%#d/%Y")
        sql_update_string = "UPDATE AFU_Child_DataCollction_Tracking " \
                            f"SET ch_child_id_{postfix}_Completed=Yes, " \
                            f"ch_child_id_{postfix}_DateCompleted = '{date_note}' " \
                            f"WHERE ch_child_id_{postfix}='{child_id}'"
        # print(sql_update_string)
        self.__cursor.execute(sql_update_string)
        self.__cursor.commit()
        return True

    def __setHouseComplete(self, house_id):
        date_note = date.today().strftime("%#m/%#d/%Y")

        sql_update_string = f"UPDATE AFU_House SET updateCompleted=Yes, date_completed='{date_note}', proc_status='NEW', date_imported_to_ASISt=null " \
                            f"WHERE ID={house_id}"
        # print(sql_update_string)
        self.__cursor.execute(sql_update_string)
        self.__cursor.commit()
        return True

    def __updateHouseAddress(self, house_id, new_address):
        sql_update_string = f"UPDATE AFU_House SET address='{new_address}' WHERE ID={house_id}"
        # print(sql_update_string)
        self.__cursor.execute(sql_update_string)
        self.__cursor.commit()
        return True

    def __updateCollectedDateManualAfu(self):
        date_entered = date.today().strftime("%#m/%#d/%Y")
        wb = openpyxl.load_workbook(self.__manual_afu_path)
        sheet = wb.active

        all_child_ids = extractChildIdsFrom(self.__matched_profiles)

        row = 3
        for rows in sheet.iter_rows(min_row=3, max_row=9143, min_col=2, max_col=2):
            for cell in rows:
                if cell.value in all_child_ids:
                    if sheet.cell(row=row, column=14).value is None:
                        sheet.cell(row=row, column=14).value = date_entered
            row += 1
        wb.save(self.__manual_afu_path)

    def __renameUnmatchedProfilesSchoolName(self):
        not_found_ids = extractChildIdsFrom(self.__non_matched_profiles)
        wb = openpyxl.load_workbook(self.__manual_afu_path)
        sheet = wb.active
        row = 3
        for rows in sheet.iter_rows(min_row=3, max_row=9143, min_col=2, max_col=2):
            child_id = str(sheet.cell(row=row, column=2).value).strip()
            if child_id in not_found_ids:
                new_school = getCommunity(child_id) + " - " + getSchool(child_id)
                print(child_id + " " + new_school)
                sheet.cell(row=row, column=1).value = new_school
            row += 1
        wb.save(self.__manual_afu_path)

    def __markedNoPhotoRows(self):
        no_photo_ids = extractChildIdsFrom(self.__has_no_photo_profiles)
        wb = openpyxl.load_workbook(self.__manual_afu_path)

        sheet = wb.active

        fill_color = '00FFFF00'
        has_photo_count = 0
        id_col = 3
        start_row = 3
        for rows in sheet.iter_rows(min_row=start_row, max_row=9143, min_col=id_col, max_col=id_col):
            for cell in rows:
                if str(cell.value).strip() in no_photo_ids:
                    has_photo_count += 1
                    fill_with = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    if cell.fill != fill_with:
                        cell.fill = fill_with

        wb.save(self.__manual_afu_path)

        print(f"Found {bcolors.WARNING}{has_photo_count}{bcolors.ENDC} photo file match id in excel!")
        print(f"Done colorizing rows which has photo")

    def __getProfileDataBySchool(self):
        return getManualAfuDbFormat(self.__manual_afu_path, self.__school_name)

    def __validateProfileData(self):
        # PERFORM VALIDATION ON ALL FILLED ATTRIBUTE ON MANUAL AFU
        # IN ORDER TO MAKE SURE ALL CODES FOLLOW THE RULE
        return profileDataAreValid(self.__db_format_profiles, self.__cursor)

    def __getChildIdsFromPhotos(self):
        return getChildIdsFromPath(self.DEFAULT_PHOTO_PATH + self.__school_name, "jpg")

    def __splitExistNonExistNoPhoto(self):
        non_existed_rows = []
        existed_rows = [[], [], [], [], [], [], [], []]
        has_photo_child_ids = self.__getChildIdsFromPhotos()
        has_no_photo_rows = []

        all_children = getAllChildren(self.__cursor)
        # print("all children:", all_children)

        for row in self.__db_format_profiles:
            child_id = row[0]
            child_has_photo = child_id in has_photo_child_ids
            idx = 0
            found_in_group_and_has_photo = False
            for suffixed_group in all_children:
                clean_suffixed_group = [item[0] for item in suffixed_group]
                if (child_id in clean_suffixed_group) and child_has_photo:
                    existed_rows[idx].append(row)
                    found_in_group_and_has_photo = True
                idx += 1
            if not found_in_group_and_has_photo:
                if not child_has_photo:
                    has_no_photo_rows.append(row)
                    print(f"Id {child_id} has no photo!")
                else:
                    print(f"Id {child_id} can't be found on db!")
                    non_existed_rows.append(row)

        return existed_rows, non_existed_rows, has_no_photo_rows


